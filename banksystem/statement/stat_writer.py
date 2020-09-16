import openpyxl
import xlrd
import xlsxwriter
from statement import get_date

def statement_writer(acc_no,prev_bal,add_with,up_bal,type_transaction):
    date = get_date.date_today()
    stat = xlrd.open_workbook('account\{}.xlsx'.format(acc_no))
    stat_sheet = stat.sheet_by_name('record')
    row_val = int(stat_sheet.cell_value(0,0))


    stat_acc = openpyxl.load_workbook('account\{}.xlsx'.format(acc_no))
    stat_worksheet = stat_acc['statement']
    rec_worksheet = stat_acc['record']
    rec_val = rec_worksheet.cell(1,1)

    credit_debit = stat_worksheet.cell(row_val,int(type_transaction))
    credit_debit.value = add_with
    else_cell = stat_worksheet.cell(row_val,5-int(type_transaction))
    else_cell.value = '-'
    date_col = stat_worksheet.cell(row_val,1)
    date_col.value = date
    bal_cell = stat_worksheet.cell(row_val,4)
    bal_cell.value = up_bal
    rec_val.value = row_val+1
    stat_acc.save('account\{}.xlsx'.format(acc_no))

